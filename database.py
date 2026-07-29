"""Persistência SQLite do Leads Maps v2, com migrações aditivas e backups seguros."""
import csv
import os
import re
import secrets
import sqlite3
import threading
import time
import unicodedata
import urllib.parse
from datetime import datetime, timezone

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "leads.db")
FUNNEL_STATUSES = (
    "novo", "para_contatar", "contatado", "respondeu", "interessado",
    "reuniao", "cliente", "descartado",
)

# Mantido com 15 colunas para compatibilidade com instalações e integrações legadas.
SCHEMA = """
CREATE TABLE IF NOT EXISTS leads (
    place_id TEXT PRIMARY KEY, name TEXT, address TEXT, phone TEXT, email TEXT,
    email_fonte TEXT, email_sugerido TEXT, website TEXT, category TEXT, rating REAL,
    ratings_total INTEGER, latitude REAL, longitude REAL, search_query TEXT, created_at TEXT
);
"""
EXPECTED_COLUMNS = {
    "place_id": "TEXT", "name": "TEXT", "address": "TEXT", "phone": "TEXT",
    "email": "TEXT", "email_fonte": "TEXT", "email_sugerido": "TEXT",
    "website": "TEXT", "category": "TEXT", "rating": "REAL",
    "ratings_total": "INTEGER", "latitude": "REAL", "longitude": "REAL",
    "search_query": "TEXT", "created_at": "TEXT", "updated_at": "TEXT",
    "segment": "TEXT", "city": "TEXT", "uf": "TEXT", "email_quality": "TEXT",
    "email_confidence": "INTEGER", "email_domain_aligned": "INTEGER DEFAULT 0",
    "email_mx_valid": "INTEGER", "status": "TEXT DEFAULT 'novo'", "notes": "TEXT",
    "assignee": "TEXT", "tags": "TEXT", "last_contact_at": "TEXT",
    "next_followup_at": "TEXT", "exported_at": "TEXT", "normalized_email": "TEXT",
    "normalized_phone": "TEXT", "normalized_domain": "TEXT",
    "normalized_name_address": "TEXT",
    "brevo_contact_id": "TEXT", "brevo_list_id": "INTEGER", "brevo_synced_at": "TEXT",
    "brevo_last_attempt_at": "TEXT", "brevo_sync_error": "TEXT",
}
_SCHEMA_LOCK = threading.Lock()
BACKUP_FILENAME_RE = re.compile(
    r"leads-\d{8}-\d{6}-\d{6}(?:-\d+-[0-9a-f]{8})?\.db"
)


def get_db_path():
    return os.path.abspath(os.environ.get("LEADS_DB_PATH", DB_PATH))


def _plain(value):
    text = unicodedata.normalize("NFKD", str(value or ""))
    return " ".join("".join(c for c in text if not unicodedata.combining(c)).lower().split())


def normalize_email(value):
    return str(value or "").strip().lower()


def normalize_phone(value):
    digits = re.sub(r"\D", "", str(value or ""))
    return digits[-11:] if len(digits) >= 10 else ""


def normalize_domain(value):
    raw = str(value or "").strip().lower()
    if "@" in raw and "://" not in raw:
        raw = raw.rsplit("@", 1)[1]
    parsed = urllib.parse.urlsplit(raw if "://" in raw else "//" + raw)
    return (parsed.hostname or "").removeprefix("www.")


def normalize_name_address(name, address):
    return f"{_plain(name)}|{_plain(address)}" if name and address else ""


def _migrate(conn):
    # Serializa migrações entre threads/processos antes de observar colunas ausentes.
    conn.execute("BEGIN IMMEDIATE")
    try:
        existing = {row[1] for row in conn.execute("PRAGMA table_info(leads)")}
        for column, column_type in EXPECTED_COLUMNS.items():
            if column not in existing:
                # Tipos vêm de constante interna, nunca de entrada do usuário.
                conn.execute(f"ALTER TABLE leads ADD COLUMN {column} {column_type}")
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS api_usage (
                day TEXT NOT NULL, endpoint TEXT NOT NULL, requests INTEGER NOT NULL DEFAULT 0,
                units INTEGER NOT NULL DEFAULT 0, estimated_cost REAL NOT NULL DEFAULT 0,
                PRIMARY KEY(day, endpoint)
            );
            CREATE TABLE IF NOT EXISTS backups (
                filename TEXT PRIMARY KEY, created_at TEXT NOT NULL, size_bytes INTEGER NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_leads_email ON leads(normalized_email);
            CREATE INDEX IF NOT EXISTS idx_leads_phone ON leads(normalized_phone);
            CREATE INDEX IF NOT EXISTS idx_leads_domain ON leads(normalized_domain);
            CREATE INDEX IF NOT EXISTS idx_leads_name_address ON leads(normalized_name_address);
            CREATE INDEX IF NOT EXISTS idx_leads_created ON leads(created_at);
        """)
        rows = conn.execute(
            "SELECT place_id,name,address,phone,email,website FROM leads WHERE "
            "normalized_email IS NULL OR normalized_phone IS NULL OR "
            "normalized_domain IS NULL OR normalized_name_address IS NULL"
        ).fetchall()
        for pid, name, address, phone, email, website in rows:
            conn.execute(
                "UPDATE leads SET normalized_email=?,normalized_phone=?,normalized_domain=?,"
                "normalized_name_address=?,status=COALESCE(NULLIF(status,''),'novo') WHERE place_id=?",
                (normalize_email(email), normalize_phone(phone), normalize_domain(website or email),
                 normalize_name_address(name, address), pid),
            )
        conn.commit()
    except sqlite3.Error:
        conn.rollback()
        raise


def get_connection():
    path = get_db_path()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    conn = sqlite3.connect(path, timeout=5)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA busy_timeout = 5000")
    conn.execute("PRAGMA foreign_keys = ON")
    with _SCHEMA_LOCK:
        conn.execute("PRAGMA journal_mode = WAL")
        conn.execute(SCHEMA)
        existing = {row[1] for row in conn.execute("PRAGMA table_info(leads)")}
        if set(EXPECTED_COLUMNS) - existing and conn.execute("SELECT EXISTS(SELECT 1 FROM leads)").fetchone()[0]:
            create_backup(conn, retention=30)
        _migrate(conn)
    return conn


def _lead_values(lead):
    values = dict(lead)
    values.setdefault("created_at", datetime.now(timezone.utc).isoformat())
    values["updated_at"] = datetime.now(timezone.utc).isoformat()
    for key in EXPECTED_COLUMNS:
        values.setdefault(key, None)
    values["status"] = values.get("status") or "novo"
    values["normalized_email"] = normalize_email(values.get("email"))
    values["normalized_phone"] = normalize_phone(values.get("phone"))
    values["normalized_domain"] = normalize_domain(values.get("website") or values.get("email"))
    values["normalized_name_address"] = normalize_name_address(values.get("name"), values.get("address"))
    return values


def _find_duplicate(conn, values):
    clauses, params = ["place_id=?"], [values["place_id"]]
    for field in ("normalized_email", "normalized_phone", "normalized_name_address"):
        if values[field]:
            clauses.append(f"{field}=?")
            params.append(values[field])
    duplicate = conn.execute(
        f"SELECT * FROM leads WHERE {' OR '.join(clauses)} ORDER BY CASE WHEN place_id=? THEN 0 ELSE 1 END LIMIT 1",
        params + [values["place_id"]],
    ).fetchone()
    if duplicate is not None or not values["normalized_domain"]:
        return duplicate
    # Domínio é apenas um sinal auxiliar: hospedagens/redes sociais são compartilhadas.
    # Só confirma duplicidade quando o nome comercial normalizado também coincide.
    for row in conn.execute("SELECT * FROM leads WHERE normalized_domain=?", (values["normalized_domain"],)):
        if _plain(row["name"]) and _plain(row["name"]) == _plain(values.get("name")):
            return row
    return None


def upsert_lead(conn, lead: dict):
    """Insere/mescla por chaves normalizadas. Retorna metadados para contar o alvo."""
    values = _lead_values(lead)
    existing = _find_duplicate(conn, values)
    columns = list(EXPECTED_COLUMNS)
    if existing is None:
        sql = f"INSERT INTO leads ({','.join(columns)}) VALUES ({','.join(':'+c for c in columns)})"
        conn.execute(sql, values)
        conn.commit()
        return {"is_new": True, "gained_email": bool(values["email"]), "place_id": values["place_id"]}

    old = dict(existing)
    gained = not bool(old.get("email")) and bool(values.get("email"))
    old_confidence = old.get("email_confidence") or 0
    incoming_confidence = values.get("email_confidence") or 0
    if old.get("email") and (not values.get("email") or incoming_confidence <= old_confidence):
        for field in (
            "email", "email_fonte", "email_sugerido", "email_quality", "email_confidence",
            "email_domain_aligned", "email_mx_valid", "normalized_email",
        ):
            values[field] = old.get(field)
    # Não rebaixa dados: campos vazios da busca nunca apagam valores existentes/CRM.
    mutable = [
        "name", "address", "phone", "email", "email_fonte", "email_sugerido", "website",
        "category", "rating", "ratings_total", "latitude", "longitude", "search_query",
        "segment", "city", "uf", "email_quality", "email_confidence", "email_domain_aligned",
        "email_mx_valid", "normalized_email", "normalized_phone", "normalized_domain",
        "normalized_name_address", "updated_at",
    ]
    merged = {field: values.get(field) if values.get(field) not in (None, "") else old.get(field) for field in mutable}
    merged["pid"] = old["place_id"]
    conn.execute(
        "UPDATE leads SET " + ",".join(f"{field}=:{field}" for field in mutable) + " WHERE place_id=:pid",
        merged,
    )
    conn.commit()
    return {"is_new": False, "gained_email": gained, "place_id": old["place_id"]}


def count_leads(conn, filtro=None):
    where = ""
    if filtro == "com_email": where = " WHERE COALESCE(email,'')!=''"
    elif filtro == "sem_email": where = " WHERE COALESCE(email,'')=''"
    return conn.execute("SELECT COUNT(*) FROM leads" + where).fetchone()[0]


def build_lead_filters(filters=None):
    filters = filters or {}
    clauses, params = ["1=1"], []
    mappings = {"quality": "email_quality", "segment": "segment", "city": "city", "uf": "uf", "status": "status"}
    email_filter = filters.get("email") or filters.get("filtro")
    if email_filter == "com_email": clauses.append("COALESCE(email,'')!=''")
    elif email_filter == "sem_email": clauses.append("COALESCE(email,'')='' ")
    for key, column in mappings.items():
        if filters.get(key): clauses.append(f"{column}=?"); params.append(filters[key])
    for key, column in (("phone", "phone"), ("site", "website")):
        if filters.get(key) == "sim": clauses.append(f"COALESCE({column},'')!=''")
        elif filters.get(key) == "nao": clauses.append(f"COALESCE({column},'')='' ")
    if filters.get("rating_min"):
        clauses.append("rating>=?"); params.append(float(filters["rating_min"]))
    exported = filters.get("exported")
    if exported == "sim": clauses.append("exported_at IS NOT NULL")
    elif exported == "nao": clauses.append("exported_at IS NULL")
    if filters.get("date_from"): clauses.append("created_at>=?"); params.append(filters["date_from"])
    if filters.get("date_to"): clauses.append("created_at<?"); params.append(filters["date_to"] + "T23:59:59.999999")
    if filters.get("followup") == "atrasado": clauses.append("next_followup_at < ?"); params.append(datetime.now(timezone.utc).isoformat())
    if filters.get("followup") == "agendado": clauses.append("next_followup_at IS NOT NULL")
    busca = filters.get("busca")
    if busca:
        clauses.append("(search_query LIKE ? OR name LIKE ? OR address LIKE ? OR email LIKE ?)")
        params.extend([f"%{busca}%"] * 4)
    ids = filters.get("ids")
    if ids:
        clean_ids = [str(i) for i in ids if str(i)]
        if clean_ids:
            clauses.append(f"place_id IN ({','.join('?' for _ in clean_ids)})"); params.extend(clean_ids)
    return " AND ".join(clauses), params


def query_lead_records(conn, filters=None):
    where, params = build_lead_filters(filters)
    return conn.execute(
        f"SELECT * FROM leads WHERE {where} ORDER BY CASE WHEN COALESCE(email,'')!='' THEN 0 ELSE 1 END, created_at DESC",
        params,
    ).fetchall()


def query_leads(conn, filtro=None, busca=None):
    sql = "SELECT name,address,phone,email,website,category,rating,search_query FROM leads WHERE 1=1"
    params = []
    if filtro == "com_email": sql += " AND COALESCE(email,'')!=''"
    elif filtro == "sem_email": sql += " AND COALESCE(email,'')=''"
    if busca: sql += " AND search_query LIKE ?"; params.append(f"%{busca}%")
    sql += " ORDER BY CASE WHEN COALESCE(email,'')!='' THEN 0 ELSE 1 END, created_at DESC"
    return [tuple(row) for row in conn.execute(sql, params).fetchall()]


def update_lead_crm(conn, place_id, fields):
    allowed = {"status", "notes", "assignee", "tags", "last_contact_at", "next_followup_at"}
    values = {k: fields.get(k) or None for k in allowed if k in fields}
    if "status" in values and values["status"] not in FUNNEL_STATUSES:
        raise ValueError("Status inválido")
    if not values: return False
    values["place_id"] = place_id
    cur = conn.execute("UPDATE leads SET " + ",".join(f"{k}=:{k}" for k in values if k != "place_id") + " WHERE place_id=:place_id", values)
    conn.commit()
    return cur.rowcount == 1


def record_brevo_sync(conn, place_id, contact_id, list_id, error=None, attempted_at=None):
    attempted_at = attempted_at or datetime.now(timezone.utc).isoformat()
    if error:
        cur = conn.execute(
            "UPDATE leads SET brevo_last_attempt_at=?,brevo_sync_error=? WHERE place_id=?",
            (attempted_at, str(error)[:500], place_id),
        )
    else:
        cur = conn.execute(
            "UPDATE leads SET brevo_contact_id=?,brevo_list_id=?,brevo_synced_at=?,"
            "brevo_last_attempt_at=?,brevo_sync_error=NULL WHERE place_id=?",
            (str(contact_id) if contact_id is not None else None, int(list_id), attempted_at, attempted_at, place_id),
        )
    conn.commit()
    return cur.rowcount == 1


def safe_cell(value):
    if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@", "\t", "\r")):
        return "'" + value
    return value


def _export_rows(conn, filtro=None, busca=None, filters=None):
    combined = dict(filters or {})
    if filtro: combined["filtro"] = filtro
    if busca: combined["busca"] = busca
    return query_lead_records(conn, combined)


def export_csv(conn, filepath="leads_export.csv", filtro=None, busca=None, filters=None):
    rows = _export_rows(conn, filtro, busca, filters)
    existing = [row[1] for row in conn.execute("PRAGMA table_info(leads)")]
    headers = [header for header in EXPECTED_COLUMNS if header in existing]
    with open(filepath, "w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle); writer.writerow(headers)
        for row in rows:
            if isinstance(row, sqlite3.Row):
                values = [safe_cell(row[header]) for header in headers]
            else:
                by_name = dict(zip(existing, row))
                values = [safe_cell(by_name.get(header)) for header in headers]
            writer.writerow(values)
    return filepath


def export_xlsx(conn, filepath, filters=None):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    rows = _export_rows(conn, filters=filters)
    headers = list(EXPECTED_COLUMNS)
    wb = Workbook(); ws = wb.active; ws.title = "Leads"
    ws.append(headers)
    for row in rows: ws.append([safe_cell(row[h]) for h in headers])
    fill = PatternFill("solid", fgColor="312E81")
    for cell in ws[1]: cell.font = Font(name="Arial", bold=True, color="FFFFFF"); cell.fill = fill
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    for column, header in enumerate(headers, 1):
        width = min(45, max(len(header) + 2, *(len(str(ws.cell(r, column).value or "")) + 2 for r in range(2, ws.max_row + 1))))
        ws.column_dimensions[get_column_letter(column)].width = width
        for cell in ws[get_column_letter(column)]: cell.font = Font(name="Arial", size=10); cell.alignment = Alignment(vertical="top")
    wb.save(filepath)
    return filepath


def mark_exported(conn, place_ids):
    ids = [str(value) for value in place_ids if str(value)]
    if not ids: return 0
    now = datetime.now(timezone.utc).isoformat()
    cur = conn.execute(f"UPDATE leads SET exported_at=? WHERE place_id IN ({','.join('?' for _ in ids)})", [now, *ids])
    conn.commit(); return cur.rowcount


def record_api_call(conn, endpoint, units=1, rate=0.0, ceiling=None, day=None):
    day = day or datetime.now(timezone.utc).date().isoformat()
    conn.execute("BEGIN IMMEDIATE")
    try:
        used = conn.execute(
            "SELECT COALESCE(SUM(requests),0) FROM api_usage "
            "WHERE day=? AND endpoint IN ('text_search','place_details','details')",
            (day,),
        ).fetchone()[0]
        if ceiling is not None and used >= ceiling:
            conn.rollback()
            return False
        conn.execute(
            "INSERT INTO api_usage(day,endpoint,requests,units,estimated_cost) VALUES(?,?,1,?,?) "
            "ON CONFLICT(day,endpoint) DO UPDATE SET requests=requests+1,units=units+excluded.units,estimated_cost=estimated_cost+excluded.estimated_cost",
            (day, endpoint, units, units * rate),
        )
        conn.commit()
        return True
    except sqlite3.Error:
        conn.rollback()
        raise


def api_usage_summary(conn, day=None):
    day = day or datetime.now(timezone.utc).date().isoformat()
    rows = conn.execute("SELECT endpoint,requests,units,estimated_cost FROM api_usage WHERE day=? ORDER BY endpoint", (day,)).fetchall()
    api_requests = sum(r[1] for r in rows if r[0] in {"text_search", "place_details", "details"})
    return {"day": day, "requests": sum(r[1] for r in rows), "api_requests": api_requests, "units": sum(r[2] for r in rows), "estimated_cost": sum(r[3] for r in rows), "rows": rows}


def _replace_with_retry(source, target, attempts=5):
    for attempt in range(attempts):
        try:
            os.replace(source, target)
            return
        except PermissionError:
            if attempt == attempts - 1:
                raise
            time.sleep(0.05 * (attempt + 1))


def _remove_stale_backup(path, attempts=5):
    for attempt in range(attempts):
        try:
            os.unlink(path)
            return
        except FileNotFoundError:
            return
        except PermissionError:
            if attempt == attempts - 1:
                return
            time.sleep(0.05 * (attempt + 1))


def create_backup(conn=None, retention=30, now=None):
    source = get_db_path(); folder = os.path.join(os.path.dirname(source), "backups")
    os.makedirs(folder, exist_ok=True)
    now = now or datetime.now(timezone.utc)
    filename = f"leads-{now.strftime('%Y%m%d-%H%M%S-%f')}-{os.getpid()}-{secrets.token_hex(4)}.db"
    target = os.path.join(folder, filename)
    temporary = target + ".tmp"
    source_conn = conn or get_connection()
    destination = None
    try:
        destination = sqlite3.connect(temporary)
        source_conn.backup(destination)
        integrity = destination.execute("PRAGMA integrity_check").fetchone()[0]
        if integrity != "ok":
            raise sqlite3.DatabaseError(f"Backup inválido: {integrity}")
        destination.close()
        destination = None
        _replace_with_retry(temporary, target)
    except (OSError, sqlite3.Error):
        if destination is not None:
            destination.close()
        try: os.unlink(temporary)
        except FileNotFoundError: pass
        raise
    finally:
        if conn is None: source_conn.close()
    files = sorted((f for f in os.listdir(folder) if BACKUP_FILENAME_RE.fullmatch(f)), reverse=True)
    for stale in files[max(1, retention):]:
        _remove_stale_backup(os.path.join(folder, stale))
    has_backup_table = conn is not None and conn.execute(
        "SELECT EXISTS(SELECT 1 FROM sqlite_master WHERE type='table' AND name='backups')"
    ).fetchone()[0]
    if has_backup_table:
        try:
            size_bytes = os.path.getsize(target)
        except FileNotFoundError:
            pass
        else:
            conn.execute(
                "INSERT OR REPLACE INTO backups VALUES(?,?,?)",
                (filename, now.isoformat(), size_bytes),
            )
            conn.commit()
    return target


def maybe_daily_backup(conn):
    today = datetime.now(timezone.utc).date().isoformat()
    latest = conn.execute("SELECT created_at FROM backups ORDER BY created_at DESC LIMIT 1").fetchone()
    if latest and str(latest[0]).startswith(today): return None
    return create_backup(conn)


def list_backups():
    folder = os.path.join(os.path.dirname(get_db_path()), "backups")
    if not os.path.isdir(folder): return []
    backups = []
    for filename in sorted(os.listdir(folder), reverse=True):
        if not BACKUP_FILENAME_RE.fullmatch(filename):
            continue
        try:
            size_bytes = os.path.getsize(os.path.join(folder, filename))
        except (FileNotFoundError, PermissionError):
            continue
        backups.append({"filename": filename, "size_bytes": size_bytes})
    return backups


def limpar_emails(conn, import_urllib=None):
    updates = 0
    for place_id, email in conn.execute("SELECT place_id,email FROM leads"):
        if email:
            clean = urllib.parse.unquote(email).strip()
            if clean != email:
                conn.execute("UPDATE leads SET email=?,normalized_email=? WHERE place_id=?", (clean, normalize_email(clean), place_id)); updates += 1
    conn.commit(); return updates


def apagar_todos(conn):
    conn.execute("DELETE FROM leads"); conn.commit()
