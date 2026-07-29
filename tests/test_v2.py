import csv
import os
import re
import sqlite3
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timezone

import pytest

import app as webapp
import database
import lead_quality


def connect_tmp(monkeypatch, tmp_path):
    monkeypatch.setenv("LEADS_DB_PATH", str(tmp_path / "leads.db"))
    return database.get_connection()


def lead(place_id="p1", **changes):
    value = {
        "place_id": place_id, "name": "Empresa", "address": "Rua 1",
        "phone": "+55 (84) 99999-0000", "email": None, "email_fonte": None,
        "email_sugerido": None, "website": "https://empresa.com.br", "category": "lawyer",
        "rating": 4.5, "ratings_total": 10, "latitude": -5.0, "longitude": -37.0,
        "search_query": "advocacia Mossoró RN", "created_at": datetime.now(timezone.utc).isoformat(),
        "segment": "advocacia", "city": "Mossoró", "uf": "RN",
    }
    value.update(changes)
    return value


def csrf(client, path="/"):
    html = client.get(path).get_data(as_text=True)
    return re.search(r'name="csrf_token" value="([^"]+)"', html).group(1)


def test_additive_migration_adds_v2_fields_and_operational_tables(monkeypatch, tmp_path):
    db = tmp_path / "leads.db"
    conn = sqlite3.connect(db)
    conn.execute("CREATE TABLE leads (place_id TEXT PRIMARY KEY, name TEXT, email TEXT)")
    conn.execute("INSERT INTO leads VALUES ('old', 'Legado', 'old@example.org')")
    conn.commit(); conn.close()
    monkeypatch.setenv("LEADS_DB_PATH", str(db))

    conn = database.get_connection()
    columns = {row[1] for row in conn.execute("PRAGMA table_info(leads)")}
    tables = {row[0] for row in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")}

    assert {"email_quality", "email_confidence", "status", "notes", "assignee", "tags",
            "last_contact_at", "next_followup_at", "exported_at", "segment", "city", "uf",
            "normalized_email", "normalized_phone", "normalized_domain", "normalized_name_address",
            "brevo_contact_id", "brevo_list_id", "brevo_synced_email", "brevo_synced_at", "brevo_last_attempt_at",
            "brevo_sync_error"} <= columns
    assert {"api_usage", "backups", "brevo_sync_state"} <= tables
    assert conn.execute("SELECT name FROM leads WHERE place_id='old'").fetchone()[0] == "Legado"
    exported = tmp_path / "legacy.csv"
    database.export_csv(conn, str(exported))
    with exported.open(encoding="utf-8-sig", newline="") as handle:
        row = next(csv.DictReader(handle))
    assert row["place_id"] == "old"
    assert row["name"] == "Legado"
    assert row["email"] == "old@example.org"
    migration_backups = database.list_backups()
    assert len(migration_backups) == 1
    backup_conn = sqlite3.connect(tmp_path / "backups" / migration_backups[0]["filename"])
    backup_columns = {item[1] for item in backup_conn.execute("PRAGMA table_info(leads)")}
    assert backup_columns == {"place_id", "name", "email"}
    assert backup_conn.execute("SELECT name FROM leads WHERE place_id='old'").fetchone()[0] == "Legado"
    backup_conn.close()
    conn.close()


def test_migration_backfills_brevo_success_history_for_legacy_synced_leads(monkeypatch, tmp_path):
    db = tmp_path / "legacy-brevo.db"
    conn = sqlite3.connect(db)
    conn.execute(
        "CREATE TABLE leads (place_id TEXT PRIMARY KEY,email TEXT,normalized_email TEXT,"
        "brevo_contact_id TEXT,brevo_list_id INTEGER,brevo_synced_at TEXT)"
    )
    conn.execute(
        "INSERT INTO leads VALUES (?,?,?,?,?,?)",
        (
            "legado-brevo", "Contato@Empresa.test", "contato@empresa.test",
            "321", 7, "2026-07-29T12:00:00+00:00",
        ),
    )
    conn.commit()
    conn.close()
    monkeypatch.setenv("LEADS_DB_PATH", str(db))

    conn = database.get_connection()
    history = conn.execute(
        "SELECT place_id,list_id,synced_email,contact_id,synced_at "
        "FROM brevo_sync_state WHERE place_id='legado-brevo'"
    ).fetchone()
    lead = conn.execute(
        "SELECT brevo_synced_email FROM leads WHERE place_id='legado-brevo'"
    ).fetchone()
    conn.close()

    assert tuple(history) == (
        "legado-brevo", 7, "contato@empresa.test", "321", "2026-07-29T12:00:00+00:00",
    )
    assert lead[0] == "contato@empresa.test"


def test_additive_migration_is_serialized_between_concurrent_connections(monkeypatch, tmp_path):
    db = tmp_path / "concurrent.db"
    conn = sqlite3.connect(db)
    conn.execute("CREATE TABLE leads (place_id TEXT PRIMARY KEY, name TEXT, email TEXT)")
    conn.execute("INSERT INTO leads VALUES ('old', 'Legado', 'old@example.org')")
    conn.commit()
    conn.close()
    monkeypatch.setenv("LEADS_DB_PATH", str(db))

    def connect_and_read():
        connection = database.get_connection()
        columns = {row[1] for row in connection.execute("PRAGMA table_info(leads)")}
        connection.close()
        return columns

    with ThreadPoolExecutor(max_workers=4) as pool:
        results = list(pool.map(lambda _index: connect_and_read(), range(4)))

    assert all("brevo_contact_id" in columns for columns in results)


def test_email_quality_scores_source_alignment_and_mx(monkeypatch):
    monkeypatch.setattr(lead_quality, "has_mx", lambda domain, timeout=1.0: domain == "empresa.com.br")
    high = lead_quality.assess_email(
        "contato@empresa.com.br", "https://empresa.com.br/contato", "https://www.empresa.com.br"
    )
    low = lead_quality.assess_email(
        "empresa@gmail.com", "sugerido", "https://empresa.com.br", check_mx=False
    )

    assert high == {"quality": "alta", "confidence": 95, "domain_aligned": True, "mx_valid": True}
    assert low["quality"] == "baixa"
    assert low["confidence"] <= 45
    assert low["domain_aligned"] is False


def test_upsert_deduplicates_normalized_identity_and_only_email_upgrade_counts(monkeypatch, tmp_path):
    conn = connect_tmp(monkeypatch, tmp_path)
    first = database.upsert_lead(conn, lead("p1"))
    duplicate = database.upsert_lead(conn, lead("p2", phone="5584999990000", website="http://www.empresa.com.br"))
    upgrade = database.upsert_lead(conn, lead("p3", email="CONTATO@EMPRESA.COM.BR", email_quality="alta"))
    repeat = database.upsert_lead(conn, lead("p4", email="contato@empresa.com.br"))

    assert first == {"is_new": True, "gained_email": False, "place_id": "p1"}
    assert duplicate["is_new"] is False and duplicate["gained_email"] is False
    assert upgrade["is_new"] is False and upgrade["gained_email"] is True
    assert repeat["is_new"] is False and repeat["gained_email"] is False
    assert database.count_leads(conn) == 1
    assert database.count_leads(conn, "com_email") == 1
    conn.close()


def test_shared_website_host_alone_never_merges_distinct_companies(monkeypatch, tmp_path):
    conn = connect_tmp(monkeypatch, tmp_path)
    first = database.upsert_lead(conn, lead(
        "p1", name="Empresa Alfa", address="Rua A", phone=None,
        website="https://instagram.com/empresa-alfa",
    ))
    second = database.upsert_lead(conn, lead(
        "p2", name="Empresa Beta", address="Rua B", phone=None,
        website="https://instagram.com/empresa-beta",
    ))
    conn.close()

    assert first["is_new"] is True
    assert second["is_new"] is True


def test_upsert_never_downgrades_existing_email_quality(monkeypatch, tmp_path):
    conn = connect_tmp(monkeypatch, tmp_path)
    database.upsert_lead(conn, lead(
        "p1", email="contato@empresa.com.br", email_quality="alta",
        email_confidence=95, email_domain_aligned=1, email_mx_valid=1,
    ))

    database.upsert_lead(conn, lead(
        "p1", email=None, email_quality=None, email_confidence=None,
        email_domain_aligned=0, email_mx_valid=None,
    ))
    database.upsert_lead(conn, lead(
        "p1", email="contato@empresa.com.br", email_quality="baixa",
        email_confidence=10, email_domain_aligned=0, email_mx_valid=0,
    ))

    row = conn.execute(
        "SELECT email,email_quality,email_confidence,email_domain_aligned,email_mx_valid "
        "FROM leads WHERE place_id='p1'"
    ).fetchone()
    conn.close()
    assert tuple(row) == ("contato@empresa.com.br", "alta", 95, 1, 1)


def test_home_v2_has_separate_location_radar_progress_and_external_assets(monkeypatch, tmp_path):
    monkeypatch.setenv("APP_PUBLIC_ACCESS", "1")
    monkeypatch.setenv("LEADS_DB_PATH", str(tmp_path / "ui.db"))
    response = webapp.app.test_client().get("/")
    html = response.get_data(as_text=True)

    assert response.status_code == 200
    assert 'name="segment"' in html and 'name="city"' in html and 'name="uf"' in html
    assert 'name="location"' in html
    assert 'name="limit" min="1"' in html
    assert 'name="limit" min="1" max=' not in html
    assert 'id="search-radar"' in html and 'id="search-progress"' in html
    assert 'id="search-cancel"' in html
    assert 'src="/static/search-v2.js"' in html
    assert 'href="/static/app.css"' in html
    assert "spinner" not in html.lower()


def test_stream_search_counts_only_new_or_upgraded_real_email_and_stops_at_target(monkeypatch, tmp_path):
    monkeypatch.setenv("APP_PUBLIC_ACCESS", "1")
    monkeypatch.setenv("LEADS_DB_PATH", str(tmp_path / "stream.db"))
    monkeypatch.setenv("API_DAILY_REQUEST_CEILING", "50")
    monkeypatch.delenv("GOOGLE_PLACES_TEXT_SEARCH_RATE", raising=False)
    monkeypatch.delenv("GOOGLE_PLACES_DETAILS_RATE", raising=False)
    webapp._reset_rate_limits()
    candidates = [
        {"place_id": p, "name": p, "types": [], "geometry": {"location": {}}}
        for p in ("dup", "fail", "none", "upgrade", "fresh", "unused")
    ]
    conn = database.get_connection()
    database.upsert_lead(conn, lead("dup", email="old@dup.example", website="https://dup.example"))
    database.upsert_lead(conn, lead("upgrade", website="https://upgrade.example"))
    conn.close()
    candidate_limits = []

    def text_search(query, max_results):
        candidate_limits.append(max_results)
        webapp.places_api._record_request(webapp.places_api.NEW_TEXT_URL)
        return candidates

    monkeypatch.setattr(webapp.places_api, "text_search", text_search)
    calls = []

    def details(pid):
        calls.append(pid)
        webapp.places_api._record_request(webapp.places_api.NEW_DETAILS_URL.format(place_id=pid))
        if pid == "fail":
            raise RuntimeError("falha simulada de detalhes")
        return {"name": pid, "website": f"https://{pid}.example", "formatted_phone_number": pid}

    monkeypatch.setattr(webapp.places_api, "place_details", details)
    monkeypatch.setattr(webapp.email_finder, "find_email", lambda site: (
        (f"contato@{site.split('//')[1]}", site, None) if "none" not in site else (None, None, None)
    ))
    monkeypatch.setattr(webapp.lead_quality, "assess_email", lambda *args, **kwargs: {
        "quality": "alta", "confidence": 90, "domain_aligned": True, "mx_valid": True
    })
    automatic_syncs = []
    monkeypatch.setattr(
        webapp,
        "_auto_sync_brevo_lead",
        lambda conn, place_id: automatic_syncs.append(place_id) or "sincronizado",
    )
    client = webapp.app.test_client(); token = csrf(client)
    response = client.post("/buscar", data={
        "segment": "advocacia", "city": "Mossoró", "uf": "RN", "location": "Centro",
        "limit": "2", "csrf_token": token,
    }, headers={"Accept": "application/x-ndjson"})

    text = response.get_data(as_text=True)
    conn = database.get_connection()
    usage = database.api_usage_summary(conn)
    usage_by_endpoint = {row["endpoint"]: row["requests"] for row in usage["rows"]}
    conn.close()
    assert response.mimetype == "application/x-ndjson"
    assert '"phase": "concluida"' in text
    assert '"phase": "buscando_candidatos"' in text
    assert '"phase": "consultando_detalhes"' in text
    assert '"phase": "verificando_site"' in text
    assert '"phase": "erro_detalhes"' in text
    assert '"new_email_leads": 2' in text
    assert calls == ["dup", "fail", "none", "upgrade", "fresh"]
    assert automatic_syncs == ["dup", "upgrade", "fresh"]
    assert candidate_limits == [20]
    assert usage_by_endpoint == {"place_details": 5, "text_search": 1, "website_check": 4}
    assert usage["estimated_cost"] == 0


def test_search_stops_automatic_brevo_attempts_after_systemic_failure(monkeypatch, tmp_path):
    monkeypatch.setenv("LEADS_DB_PATH", str(tmp_path / "brevo-circuit.db"))
    monkeypatch.setenv("API_DAILY_REQUEST_CEILING", "50")
    candidates = [
        {"place_id": place_id, "name": place_id, "types": [], "geometry": {"location": {}}}
        for place_id in ("a", "b", "c")
    ]

    def text_search(query, max_results):
        webapp.places_api._record_request(webapp.places_api.NEW_TEXT_URL)
        return candidates

    def details(place_id):
        webapp.places_api._record_request(
            webapp.places_api.NEW_DETAILS_URL.format(place_id=place_id),
        )
        return {"name": place_id, "website": f"https://{place_id}.example"}

    monkeypatch.setattr(webapp.places_api, "text_search", text_search)
    monkeypatch.setattr(webapp.places_api, "place_details", details)
    monkeypatch.setattr(
        webapp.email_finder,
        "find_email",
        lambda site: (f"contato@{site.split('//')[1]}", site, None),
    )
    monkeypatch.setattr(webapp.lead_quality, "assess_email", lambda *args, **kwargs: {
        "quality": "alta", "confidence": 95, "domain_aligned": True, "mx_valid": True,
    })
    automatic_syncs = []

    def fail_systemically(conn, place_id):
        automatic_syncs.append(place_id)
        return "falha_sistemica"

    monkeypatch.setattr(webapp, "_auto_sync_brevo_lead", fail_systemically)

    events = list(webapp._search_events("advocacia", "Natal", "RN", ["advocacia Natal"], 3))

    assert events[-1]["new_email_leads"] == 3
    assert automatic_syncs == ["a"]


def test_search_assets_stream_securely_and_support_cancel(monkeypatch, tmp_path):
    monkeypatch.setenv("APP_PUBLIC_ACCESS", "1")
    monkeypatch.setenv("LEADS_DB_PATH", str(tmp_path / "assets.db"))
    client = webapp.app.test_client()
    script = client.get("/static/search-v2.js")
    css = client.get("/static/app.css")

    assert script.status_code == css.status_code == 200
    javascript = script.get_data(as_text=True)
    stylesheet = css.get_data(as_text=True)
    assert "fetch(" in javascript and "FormData" in javascript
    assert 'Accept": "application/x-ndjson"' in javascript
    assert "AbortController" in javascript and ".abort()" in javascript
    assert "cancel.disabled = true" in javascript
    assert "is-complete" in javascript and ".progress-panel.is-complete" in stylesheet
    assert "textContent" in javascript and "innerHTML" not in javascript
    assert ".radar" in stylesheet and "@keyframes" in stylesheet
    assert ".progress-panel[hidden]" in stylesheet


def test_crm_edit_requires_csrf_and_persists_valid_funnel_fields(monkeypatch, tmp_path):
    monkeypatch.setenv("APP_PUBLIC_ACCESS", "1")
    conn = connect_tmp(monkeypatch, tmp_path)
    database.upsert_lead(conn, lead())
    conn.close()
    client = webapp.app.test_client()
    assert client.post("/leads/p1/editar", data={"status": "cliente"}).status_code == 400
    token = csrf(client, "/leads")
    response = client.post("/leads/p1/editar", data={
        "csrf_token": token, "status": "interessado", "notes": "Pediu proposta",
        "assignee": "Ana", "tags": "quente, inbound", "last_contact_at": "2026-07-29T10:00",
        "next_followup_at": "2026-07-30T09:00",
    })
    conn = database.get_connection(); row = conn.execute("SELECT * FROM leads WHERE place_id='p1'").fetchone(); conn.close()
    assert response.status_code == 302
    assert (row["status"], row["notes"], row["assignee"], row["tags"]) == (
        "interessado", "Pediu proposta", "Ana", "quente, inbound"
    )


def test_advanced_filters_reject_invalid_numeric_values(monkeypatch, tmp_path):
    monkeypatch.setenv("APP_PUBLIC_ACCESS", "1")
    monkeypatch.setenv("LEADS_DB_PATH", str(tmp_path / "invalid-filter.db"))

    client = webapp.app.test_client()
    assert client.get("/leads?rating_min=not-a-number").status_code == 400
    assert client.get("/leads?status=inventado").status_code == 400
    assert client.get("/leads?date_from=29-07-2026").status_code == 400


def test_advanced_filters_keep_email_group_first(monkeypatch, tmp_path):
    monkeypatch.setenv("APP_PUBLIC_ACCESS", "1")
    conn = connect_tmp(monkeypatch, tmp_path)
    database.upsert_lead(conn, lead("without", city="Natal", address="Rua A", website="https://without.example", phone="5584999990001", status="novo"))
    database.upsert_lead(conn, lead("with", city="Natal", address="Rua B", website="https://with.example", phone="5584999990002", email="a@with.example", email_quality="alta", status="novo"))
    database.upsert_lead(conn, lead("other", city="Recife", address="Rua C", phone="5584999990003", email="b@other.com", website="https://other.com"))
    rows = database.query_lead_records(conn, {"city": "Natal", "status": "novo"})
    conn.close()
    assert [row["place_id"] for row in rows] == ["with", "without"]
    html = webapp.app.test_client().get("/leads?uf=rn").get_data(as_text=True)
    assert "with" in html and "without" in html


def test_xlsx_export_is_formatted_formula_safe_and_post_marks_only_exported_rows(monkeypatch, tmp_path):
    from openpyxl import load_workbook

    monkeypatch.setenv("APP_PUBLIC_ACCESS", "1")
    conn = connect_tmp(monkeypatch, tmp_path)
    database.upsert_lead(conn, lead("safe", name="=HYPERLINK('bad')", email="a@empresa.com.br"))
    database.upsert_lead(conn, lead("other", address="Rua 2", phone="5584999990002", website="https://other.example", email="b@other.example"))
    conn.close()
    client = webapp.app.test_client(); token = csrf(client, "/exportar")
    response = client.post("/exportar", data={
        "csrf_token": token, "scope": "selected", "format": "xlsx", "ids": ["safe"]
    })
    target = tmp_path / "export.xlsx"; target.write_bytes(response.data)
    workbook = load_workbook(target); sheet = workbook["Leads"]
    values = list(sheet.iter_rows(values_only=True))
    conn = database.get_connection()
    exported = {row["place_id"]: row["exported_at"] for row in conn.execute("SELECT place_id,exported_at FROM leads")}
    conn.close()
    assert response.status_code == 200
    assert sheet.freeze_panes == "A2" and sheet.auto_filter.ref
    assert values[1][0] == "safe" and values[1][1].startswith("'")
    assert exported["safe"] and exported["other"] is None


def test_each_retry_is_recorded_against_api_limits(monkeypatch):
    recorded = []

    def fail(*args, **kwargs):
        raise webapp.places_api.requests.RequestException("offline")

    monkeypatch.setattr(webapp.places_api.requests, "get", fail)
    monkeypatch.setattr(webapp.places_api.time, "sleep", lambda _seconds: None)

    with pytest.raises(RuntimeError), webapp.places_api.usage_recorder(recorded.append):
        webapp.places_api._get(webapp.places_api.NEW_DETAILS_URL.format(place_id="p1"))

    assert recorded == ["place_details"] * webapp.places_api.MAX_RETRIES


def test_api_ceiling_is_persisted_and_enforced(monkeypatch, tmp_path):
    conn = connect_tmp(monkeypatch, tmp_path)
    assert database.record_api_call(conn, "text_search", units=1, rate=0.03, ceiling=2, day="2026-07-29")
    assert database.record_api_call(conn, "details", units=1, rate=0.02, ceiling=2, day="2026-07-29")
    assert not database.record_api_call(conn, "details", units=1, rate=0.02, ceiling=2, day="2026-07-29")
    summary = database.api_usage_summary(conn, "2026-07-29"); conn.close()
    assert summary["requests"] == 2
    assert summary["estimated_cost"] == pytest.approx(0.05)


def test_website_checks_are_reported_without_consuming_google_api_ceiling(monkeypatch, tmp_path):
    conn = connect_tmp(monkeypatch, tmp_path)
    for _ in range(3):
        assert database.record_api_call(conn, "website_check", ceiling=None, day="2026-07-29")
    assert database.record_api_call(conn, "text_search", ceiling=2, day="2026-07-29")
    assert database.record_api_call(conn, "place_details", ceiling=2, day="2026-07-29")
    assert not database.record_api_call(conn, "place_details", ceiling=2, day="2026-07-29")

    summary = database.api_usage_summary(conn, "2026-07-29")
    conn.close()

    assert summary["requests"] == 5
    assert summary["api_requests"] == 2


def test_failed_backup_never_exposes_partial_database(monkeypatch, tmp_path):
    monkeypatch.setenv("LEADS_DB_PATH", str(tmp_path / "leads.db"))

    class BrokenSource:
        def backup(self, _destination):
            raise sqlite3.Error("falha simulada")

    with pytest.raises(sqlite3.Error):
        database.create_backup(
            BrokenSource(), retention=30,
            now=datetime(2026, 7, 29, 11, 0, tzinfo=timezone.utc),
        )

    assert database.list_backups() == []
    assert not list((tmp_path / "backups").glob("*.tmp"))


def test_online_backups_stay_beside_database_and_retention_is_bounded(monkeypatch, tmp_path):
    conn = connect_tmp(monkeypatch, tmp_path)
    database.upsert_lead(conn, lead())
    paths = []
    for index in range(4):
        paths.append(database.create_backup(conn, retention=3, now=datetime(2026, 7, 29, 10, 0, index, tzinfo=timezone.utc)))
    listed = database.list_backups(); conn.close()
    backup_dir = tmp_path / "backups"
    assert all(os.path.dirname(path) == str(backup_dir) for path in paths)
    assert len(listed) == 3
    assert all(".." not in item["filename"] and "/" not in item["filename"] for item in listed)


def test_backup_retries_transient_windows_replace_lock(monkeypatch, tmp_path):
    conn = connect_tmp(monkeypatch, tmp_path)
    database.upsert_lead(conn, lead())
    real_replace = os.replace
    attempts = []

    def flaky_replace(source, target):
        attempts.append((source, target))
        if len(attempts) == 1:
            raise PermissionError("arquivo temporariamente bloqueado")
        return real_replace(source, target)

    monkeypatch.setattr(database.os, "replace", flaky_replace)
    path = database.create_backup(
        conn,
        now=datetime(2026, 7, 29, 12, 0, tzinfo=timezone.utc),
    )
    conn.close()

    assert len(attempts) == 2
    assert os.path.exists(path)
    assert not list((tmp_path / "backups").glob("*.tmp"))


def test_backups_created_at_same_instant_have_unique_files(monkeypatch, tmp_path):
    conn = connect_tmp(monkeypatch, tmp_path)
    database.upsert_lead(conn, lead())
    instant = datetime(2026, 7, 29, 12, 30, tzinfo=timezone.utc)

    first = database.create_backup(conn, now=instant)
    second = database.create_backup(conn, now=instant)
    conn.close()

    assert first != second
    assert os.path.exists(first)
    assert os.path.exists(second)
    assert len(database.list_backups()) == 2


def test_backup_retention_tolerates_file_removed_by_another_process(monkeypatch, tmp_path):
    conn = connect_tmp(monkeypatch, tmp_path)
    database.upsert_lead(conn, lead())
    for index in range(2):
        database.create_backup(
            conn,
            retention=2,
            now=datetime(2026, 7, 29, 13, 0, index, tzinfo=timezone.utc),
        )
    real_unlink = os.unlink
    raced = False

    def concurrent_unlink(path):
        nonlocal raced
        if not raced and path.endswith(".db"):
            raced = True
            real_unlink(path)
            raise FileNotFoundError(path)
        return real_unlink(path)

    monkeypatch.setattr(database.os, "unlink", concurrent_unlink)
    database.create_backup(
        conn,
        retention=2,
        now=datetime(2026, 7, 29, 13, 0, 2, tzinfo=timezone.utc),
    )
    conn.close()

    assert raced
    assert len(database.list_backups()) == 2


def test_backup_retention_tolerates_file_locked_by_another_process(monkeypatch, tmp_path):
    conn = connect_tmp(monkeypatch, tmp_path)
    database.upsert_lead(conn, lead())
    for index in range(2):
        database.create_backup(
            conn,
            retention=2,
            now=datetime(2026, 7, 29, 13, 30, index, tzinfo=timezone.utc),
        )
    real_unlink = os.unlink
    attempts = 0

    def locked_unlink(path):
        nonlocal attempts
        if path.endswith(".db"):
            attempts += 1
            raise PermissionError("arquivo bloqueado por outro processo")
        return real_unlink(path)

    monkeypatch.setattr(database.os, "unlink", locked_unlink)
    database.create_backup(
        conn,
        retention=2,
        now=datetime(2026, 7, 29, 13, 30, 2, tzinfo=timezone.utc),
    )
    conn.close()

    assert attempts >= 1


def test_backup_metadata_tolerates_target_removed_by_concurrent_retention(monkeypatch, tmp_path):
    conn = connect_tmp(monkeypatch, tmp_path)
    database.upsert_lead(conn, lead())
    real_getsize = os.path.getsize

    def concurrent_getsize(path):
        if path.endswith(".db"):
            raise FileNotFoundError(path)
        return real_getsize(path)

    monkeypatch.setattr(database.os.path, "getsize", concurrent_getsize)
    path = database.create_backup(
        conn,
        retention=1,
        now=datetime(2026, 7, 29, 14, 0, tzinfo=timezone.utc),
    )
    conn.close()

    assert path.endswith(".db")


def test_list_backups_tolerates_file_removed_during_size_read(monkeypatch, tmp_path):
    conn = connect_tmp(monkeypatch, tmp_path)
    database.upsert_lead(conn, lead())
    first = database.create_backup(
        conn,
        now=datetime(2026, 7, 29, 14, 30, 0, tzinfo=timezone.utc),
    )
    database.create_backup(
        conn,
        now=datetime(2026, 7, 29, 14, 30, 1, tzinfo=timezone.utc),
    )
    conn.close()
    real_getsize = os.path.getsize
    removed = False

    def concurrent_getsize(path):
        nonlocal removed
        if not removed and path == first:
            removed = True
            os.unlink(path)
            raise FileNotFoundError(path)
        return real_getsize(path)

    monkeypatch.setattr(database.os.path, "getsize", concurrent_getsize)
    listed = database.list_backups()

    assert removed
    assert len(listed) == 1
